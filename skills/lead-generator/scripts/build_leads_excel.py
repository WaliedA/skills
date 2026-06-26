#!/usr/bin/env python3
"""
Lead Generator — Excel Builder
Inbound LLC Fleet/Telematics Sales Prospecting
Usage: python build_leads_excel.py '<json_data>' '<output_path>'
"""

import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Brand Colors ───────────────────────────────────────────────────────────
INBOUND_DARK  = "1B3A5C"   # Dark navy  — header background
INBOUND_MID   = "2E6DA4"   # Mid blue   — accent row
INBOUND_LIGHT = "D6E4F0"   # Light blue — alternate row
WHITE         = "FFFFFF"
GOLD          = "F0A500"   # Accent for rating column
GREEN_OK      = "E8F5E9"
GREY_EMPTY    = "F5F5F5"

# ── Column Definitions ─────────────────────────────────────────────────────
COLUMNS = [
    ("Company Name",         28),
    ("Industry / Category",  22),
    ("Phone / WhatsApp",     20),
    ("Email",                28),
    ("Website",              28),
    ("LinkedIn Profile",     30),
    ("Physical Address",     32),
    ("City",                 14),
    ("Country",              12),
    ("Latitude",             12),
    ("Longitude",            12),
    ("Google Maps Rating",   20),
    ("No. of Reviews",       16),
    ("Source",               20),
    ("Notes",                40),
]

FIELD_MAP = [
    "company_name", "industry", "phone", "email", "website",
    "linkedin", "address", "city", "country", "latitude", "longitude",
    "gmaps_rating", "gmaps_reviews", "source", "notes"
]

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def build_excel(leads: list, output_path: str):
    wb = Workbook()

    # ── Sheet 1: Leads ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leads"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Header row
    header_font   = Font(name="Arial", bold=True, color=WHITE, size=10)
    header_fill   = PatternFill("solid", fgColor=INBOUND_DARK)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 32

    # Data rows
    alt_fill  = PatternFill("solid", fgColor=INBOUND_LIGHT)
    base_fill = PatternFill("solid", fgColor=WHITE)
    data_font = Font(name="Arial", size=9)
    link_font = Font(name="Arial", size=9, color=INBOUND_MID, underline="single")
    empty_fill = PatternFill("solid", fgColor=GREY_EMPTY)

    for row_idx, lead in enumerate(leads, start=2):
        fill = alt_fill if row_idx % 2 == 0 else base_fill
        ws.row_dimensions[row_idx].height = 18

        for col_idx, field in enumerate(FIELD_MAP, start=1):
            value = lead.get(field, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 15))

            # Hyperlink styling for URL fields
            if field in ("website", "linkedin") and value and str(value).startswith("http"):
                cell.font = link_font
                cell.hyperlink = str(value)
            elif field == "phone" and value:
                wa_url = f"https://wa.me/{str(value).replace('+','').replace(' ','')}"
                cell.font = Font(name="Arial", size=9, color="075E54", underline="single")
                cell.hyperlink = wa_url
            elif not value or str(value).strip() == "":
                cell.fill = empty_fill
                cell.font = Font(name="Arial", size=9, color="AAAAAA", italic=True)
                cell.value = "—"
            else:
                cell.font = data_font
                cell.fill = fill

    # AutoFilter for filtering/sorting (more compatible than Table)
    last_row = len(leads) + 1
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Rating color scale (column 12)
    if len(leads) > 0:
        rating_col = get_column_letter(12)
        ws.conditional_formatting.add(
            f"{rating_col}2:{rating_col}{last_row}",
            ColorScaleRule(
                start_type="num", start_value=1, start_color="FF4444",
                mid_type="num",   mid_value=3,   mid_color="FFAA00",
                end_type="num",   end_value=5,   end_color="00BB44"
            )
        )

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    title_font  = Font(name="Arial", bold=True, size=14, color=INBOUND_DARK)
    label_font  = Font(name="Arial", bold=True, size=10, color=INBOUND_DARK)
    value_font  = Font(name="Arial", size=10)
    section_fill = PatternFill("solid", fgColor=INBOUND_LIGHT)

    ws2["A1"] = "Inbound LLC — Lead Generation Report"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws2["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
    ws2.merge_cells("A2:B2")
    ws2["A2"].alignment = Alignment(horizontal="center")

    ws2.row_dimensions[3].height = 8

    stats = [
        ("Total Leads",        len(leads)),
        ("With Phone",         sum(1 for l in leads if l.get("phone"))),
        ("With Email",         sum(1 for l in leads if l.get("email"))),
        ("With Website",       sum(1 for l in leads if l.get("website"))),
        ("With LinkedIn",      sum(1 for l in leads if l.get("linkedin"))),
        ("With Coordinates",   sum(1 for l in leads if l.get("latitude") and l.get("longitude"))),
    ]

    # Source breakdown
    from collections import Counter
    sources = Counter(l.get("source", "Unknown") for l in leads)
    cities  = Counter(l.get("city",   "Unknown") for l in leads)

    row = 4
    ws2.cell(row=row, column=1, value="📊 OVERVIEW").font = label_font
    ws2.cell(row=row, column=1).fill = section_fill
    ws2.merge_cells(f"A{row}:B{row}")
    row += 1

    for label, val in stats:
        ws2.cell(row=row, column=1, value=label).font = label_font
        ws2.cell(row=row, column=2, value=val).font   = value_font
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="📡 BY SOURCE").font = label_font
    ws2.cell(row=row, column=1).fill = section_fill
    ws2.merge_cells(f"A{row}:B{row}")
    row += 1

    for src, count in sources.most_common():
        ws2.cell(row=row, column=1, value=src).font   = value_font
        ws2.cell(row=row, column=2, value=count).font = value_font
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="🏙️ BY CITY").font = label_font
    ws2.cell(row=row, column=1).fill = section_fill
    ws2.merge_cells(f"A{row}:B{row}")
    row += 1

    for city, count in cities.most_common():
        ws2.cell(row=row, column=1, value=city).font   = value_font
        ws2.cell(row=row, column=2, value=count).font  = value_font
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    # Apply borders to summary data
    for r in range(4, row):
        for c in [1, 2]:
            ws2.cell(row=r, column=c).border = thin_border()

    # ── Final save ─────────────────────────────────────────────────────────
    wb.save(output_path)
    print(json.dumps({"status": "success", "output": output_path, "leads": len(leads)}))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(json.dumps({"status": "error", "message": "Usage: script.py '<json>' '<output_path>'"}))
        sys.exit(1)

    raw_json    = sys.argv[1]
    output_path = sys.argv[2]

    try:
        leads = json.loads(raw_json)
    except json.JSONDecodeError as e:
        print(json.dumps({"status": "error", "message": f"JSON parse error: {e}"}))
        sys.exit(1)

    build_excel(leads, output_path)
